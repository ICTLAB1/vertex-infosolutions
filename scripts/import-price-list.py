#!/usr/bin/env python3
"""
Turn a Savex channel price list into the catalogue's Microsoft price data.

    pip install openpyxl
    python3 scripts/import-price-list.py ~/Downloads/SAVEX_CHANNEL_PRICE_LIST_SEPT_2026.xlsx

Writes prisma/data/microsoft-price-list.json, which the seed reads.

The workbook must never be committed. It carries `Unit Sell Price` — what we
pay the distributor — beside the public list price, and a repository is not
the place for a buy price. This script copies out the public columns only, and
refuses to emit anything derived from the cost column.

What it keeps, and why:

  Segment == "Commercial"   Education, Charity and Government pricing needs
                            eligibility we cannot verify at checkout.
  TermDuration == "P1Y"     A storefront that delivers a key sells a year, not
                            a monthly commitment it cannot enforce.
  ERP Price > 0             A zero is a trial SKU with no price to charge.

Where the same product and SKU appear under several billing plans, annual
billing wins: it is the cheaper of the two for the customer and the only one
that matches a single up-front payment.
"""

from __future__ import annotations

import json
import pathlib
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover - a setup error, not a code path
    sys.exit("openpyxl is not installed. Run: pip install openpyxl")

SHEETS = ("NCE", "SUBSCRIPION")
BILLING_PREFERENCE = ("Annual", "OneTime", "Triennial", "Monthly")
OUT = pathlib.Path(__file__).resolve().parent.parent / "prisma" / "data" / "microsoft-price-list.json"


def read(sheet) -> list[dict]:
    header = [cell.value for cell in sheet[1]]
    at = {name: i for i, name in enumerate(header)}
    # The two sheets disagree on the column name for the same number.
    erp = "ERP Price" if "ERP Price" in at else "ERP"
    for required in ("ProductId", "SkuId", "SkuTitle", "Segment", "TermDuration", "BillingPlan", erp):
        if required not in at:
            sys.exit(f"Sheet {sheet.title!r} has no {required!r} column.")

    out = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        price = row[at[erp]]
        if row[at["Segment"]] != "Commercial":
            continue
        if row[at["TermDuration"]] != "P1Y":
            continue
        if not isinstance(price, (int, float)) or price <= 0:
            continue
        out.append(
            {
                "productId": str(row[at["ProductId"]]).strip(),
                "skuId": str(row[at["SkuId"]]).strip(),
                "title": str(row[at["SkuTitle"]] or "").strip(),
                "tags": str(row[at["Tags"]] or "").strip(),
                "billing": str(row[at["BillingPlan"]] or "").strip(),
                # Paise, exclusive of GST. Microsoft publishes India list prices
                # ex-tax; the seed adds GST when it writes the shelf price.
                "listExGstMinor": round(price * 100),
            }
        )
    return out


def main() -> None:
    if len(sys.argv) != 2:
        sys.exit(f"usage: {sys.argv[0]} <price-list.xlsx>")

    book = openpyxl.load_workbook(sys.argv[1], data_only=True)
    rows: list[dict] = []
    for name in SHEETS:
        if name in book.sheetnames:
            rows.extend(read(book[name]))

    best: dict[tuple[str, str], dict] = {}
    for row in rows:
        key = (row["productId"], row["skuId"])
        rank = BILLING_PREFERENCE.index(row["billing"]) if row["billing"] in BILLING_PREFERENCE else 99
        current = best.get(key)
        if current is None or rank < current["_rank"]:
            best[key] = {**row, "_rank": rank}

    kept = sorted(
        ({k: v for k, v in row.items() if not k.startswith("_")} for row in best.values()),
        key=lambda r: (r["title"].lower(), r["productId"], r["skuId"]),
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(kept, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"{len(kept)} commercial one-year SKUs -> {OUT.relative_to(OUT.parent.parent.parent)}")


if __name__ == "__main__":
    main()
