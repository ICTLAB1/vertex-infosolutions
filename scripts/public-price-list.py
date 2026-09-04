#!/usr/bin/env python3
"""
Make a shareable copy of a distributor price list: public columns only.

    pip install openpyxl
    python3 scripts/public-price-list.py ~/Downloads/SAVEX_CHANNEL_PRICE_LIST_SEPT_2026.xlsx
    python3 scripts/public-price-list.py ~/Downloads/Channel_Pricelist.xlsx

Writes <name>-public-list-price.xlsx beside the input.

Why this exists
---------------
A distributor workbook carries two prices per row: what the publisher tells
the world a licence costs, and what we pay for it. The second one is the
business. Publishing it — to a customer, a marketplace, a contractor, an
AI assistant, anyone — hands over the margin on every line, and the ratio
between the two columns gives it away even if only one of them is labelled.

So there is no safe way to "tidy up" the original file and send it. This
builds a **new** workbook and copies across only the cells it has been told
are public. Nothing else can survive the trip: not a hidden sheet, not a
cached formula result, not a defined name, not the document properties, not
a stray column somebody added last month.

The allowlist is the point
--------------------------
Columns are kept by name from an explicit list, and an unrecognised header is
a hard error rather than a shrug. Next month's price list will have a column
this file has never seen, and the two ways that can go are "it was dropped
because nobody had classified it" and "it was published because nobody had
classified it". Only the first is survivable, and refusing outright is the
only way to be sure which one happened.
"""

from __future__ import annotations

import pathlib
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover - a setup error, not a code path
    sys.exit("openpyxl is not installed. Run: pip install openpyxl")


# The two publishers' own published figures — Microsoft's Estimated Retail
# Price and Adobe's Estimated Street Price. Those are the numbers on the shelf,
# and the only money that may leave this building.
PUBLIC = {
    # Savex / Microsoft
    "Publisher",
    "ChangeIndicator",
    "ProductId",
    "SkuId",
    "SkuTitle",
    "Tags",
    "TermDuration",
    "BillingPlan",
    "ERP Price",
    "ERP",
    # Adobe VIP
    "ACD Indicator",
    "ACD Description",
    "ACD Effective Date",
    "First Order Date",
    "Last Order Date",
    "Part Number",
    "Product Family",
    "Version",
    "Operating System",
    "Language",
    "Product Type",
    "Product Type Detail",
    "Additional Detail",
    "Users",
    "Metric",
    "Bridge",
    "Level Detail",
    "Duration",
    "Media",
    "UPC/EAN Code",
    # Adobe's Estimated Street Price — their published per-seat figure, and the
    # Adobe half of the only money this script lets out.
    "ESP per Year/Per Txn",
    # A VIP licensing attribute, not a cost: mostly zero, otherwise a band code
    # like HVD_L17_POST. It says nothing about what we paid.
    "Pool",
    "Points",
    # Ordering conditions ("Approval Required from Adobe"), not prices.
    "Remarks",
    # Shared
    "Segment",
}

# Named rather than merely absent from PUBLIC, so the error message can say
# *why* a column is being dropped instead of "unknown", and so that renaming
# one upstream trips the unrecognised-header check rather than passing quietly.
WITHHELD = {
    "Unit Sell Price": "our buy price from the distributor",
    "Discounted Price": "the buy price after any negotiated discount",
    "Discount %": "the discount off list, which is the margin",
    "Qty": "quote-builder scratch, multiplied against the buy price",
    "Total": "quantity times the buy price",
    # Adobe's distributor transfer price. Worse than Microsoft's equivalent,
    # because it sits below the street price on every single row at close to a
    # constant ratio — so publishing it gives away not one licence's margin but
    # the discount rate on the whole book.
    "DTP per Year /Per TXn": "the distributor transfer price — what we pay Adobe",
}


def norm(header: object) -> str:
    """Trailing spaces in a header are a formatting accident, not a column."""
    return str(header).strip() if header is not None else ""


def find_header(rows) -> list[str] | None:
    """
    The first row that looks like column names, consuming everything above it.

    Savex starts at row 1; Adobe leaves two blank rows above its header. A
    fixed row number would either read Adobe's blanks as the column names —
    keeping nothing, and silently producing an empty file that looks fine — or
    need a per-publisher special case for something both files agree on: the
    header is the first row with several filled cells.
    """
    for _ in range(20):
        try:
            row = next(rows)
        except StopIteration:
            return None
        cells = [norm(cell) for cell in row]
        if sum(1 for cell in cells if cell) >= 3:
            return cells
    return None


def redact(source: pathlib.Path) -> pathlib.Path:
    book = openpyxl.load_workbook(source, read_only=True, data_only=True)
    out = openpyxl.Workbook()
    out.remove(out.active)

    unknown: list[str] = []
    kept_total = 0

    for name in book.sheetnames:
        sheet = book[name]
        rows = sheet.iter_rows(values_only=True)
        header = find_header(rows)
        if header is None:
            continue  # An empty sheet has nothing to leak.

        for column in header:
            if column and column not in PUBLIC and column not in WITHHELD:
                unknown.append(f"{name}.{column}")

        keep = [i for i, column in enumerate(header) if column in PUBLIC]
        if not keep:
            continue

        page = out.create_sheet(name)
        page.append([header[i] for i in keep])
        for row in rows:
            # A row shorter than the header is a trailing blank, not data.
            page.append([row[i] if i < len(row) else None for i in keep])
            kept_total += 1

        dropped = [c for c in header if c in WITHHELD]
        print(f"  {name}: kept {len(keep)} columns, withheld {len(dropped)}")

    if unknown:
        sys.exit(
            "Refusing to write anything. These columns are not in the "
            "allowlist, so this script cannot tell whether they are public:\n"
            + "\n".join(f"  - {c}" for c in unknown)
            + "\n\nAdd each one to PUBLIC or to WITHHELD in this file, with a "
            "reason, and run it again."
        )

    target = source.with_name(f"{source.stem}-public-list-price.xlsx")
    out.save(target)
    print(f"\n{kept_total} rows written to {target}")
    return target


def main() -> None:
    if len(sys.argv) != 2:
        sys.exit(f"Usage: {sys.argv[0]} <price-list.xlsx>")
    source = pathlib.Path(sys.argv[1]).expanduser()
    if not source.exists():
        sys.exit(f"No such file: {source}")

    print(f"Reading {source.name}")
    print("Withholding: " + ", ".join(sorted(WITHHELD)))
    redact(source)


if __name__ == "__main__":
    main()
