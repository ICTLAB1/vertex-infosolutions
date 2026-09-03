#!/usr/bin/env python3
"""
Turn an Adobe VIP channel price list into the catalogue's Adobe price data.

    pip install openpyxl
    python3 scripts/import-adobe-price-list.py ~/Downloads/Channel_Pricelist.xlsx

Writes prisma/data/adobe-price-list.json, which the seed reads.

As with the Microsoft list, the workbook must never be committed. It carries
`DTP per Year /Per TXn` — the distributor transfer price, what we pay — beside
the street price. This script copies out the public columns only.

What it keeps, and why:

  Segment == "Commercial"     Education pricing needs eligibility we cannot
                              verify at checkout.
  Duration == "12 Months"     A storefront sells a year.
  Users == "1 User"           Per-seat licences, which is what a basket sells.
                              Credit packs and per-transaction items are priced
                              per pack, not per seat, and would multiply wrongly
                              against a quantity box.
  Level 1                     Adobe prices in volume bands — 1-9, 10-49, 50-99,
                              100+ — and the shelf price has to be the one a
                              buyer of one seat pays. The cheaper bands are real
                              and are offered by quotation; see the note the
                              seed puts on every Adobe listing.
  ESP > 0                     A zero is a placeholder row.
"""

from __future__ import annotations

import json
import pathlib
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is not installed. Run: pip install openpyxl")

SHEET = "Commercial"
HEADER_ROW = 3
OUT = pathlib.Path(__file__).resolve().parent.parent / "prisma" / "data" / "adobe-price-list.json"


def main() -> None:
    if len(sys.argv) != 2:
        sys.exit(f"usage: {sys.argv[0]} <price-list.xlsx>")

    book = openpyxl.load_workbook(sys.argv[1], data_only=True)
    if SHEET not in book.sheetnames:
        sys.exit(f"No {SHEET!r} sheet. Found: {', '.join(book.sheetnames)}")
    sheet = book[SHEET]

    header = [cell.value for cell in sheet[HEADER_ROW]]
    at = {name: i for i, name in enumerate(header) if name}
    for required in ("Part Number", "Product Family", "Duration", "Users",
                     "Level Detail", "ESP per Year/Per Txn"):
        if required not in at:
            sys.exit(f"Sheet {SHEET!r} has no {required!r} column.")

    best: dict[str, dict] = {}
    for row in sheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        part = row[at["Part Number"]]
        price = row[at["ESP per Year/Per Txn"]]
        if not part:
            continue
        if str(row[at["Duration"]]) != "12 Months":
            continue
        if str(row[at["Users"]]) != "1 User":
            continue
        if not str(row[at["Level Detail"]]).startswith("Level 1 1 - 9"):
            continue
        if not isinstance(price, (int, float)) or price <= 0:
            continue

        family = str(row[at["Product Family"]]).strip()
        # One row per product family. Several part numbers can share a family
        # where Adobe splits by platform or licence type; the price is the same,
        # and a shop needs one listing rather than three identical ones.
        if family in best:
            continue
        best[family] = {
            "partNumber": str(part).strip(),
            "family": family,
            "productType": str(row[at["Product Type"]] or "").strip(),
            # Paise, exclusive of GST. Adobe publishes India street prices
            # ex-tax; the seed adds GST when it writes the shelf price.
            "listExGstMinor": round(price * 100),
        }

    kept = sorted(best.values(), key=lambda r: r["family"].lower())
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(kept, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"{len(kept)} commercial one-year Adobe SKUs -> {OUT.name}")


if __name__ == "__main__":
    main()
